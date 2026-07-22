"""Uploaded file extraction and analysis helpers for the runtime.

Boundary decision: keep this module as the compatibility surface for runtime
file analysis until a split is needed. See docs/file-analysis-boundary.md.
"""

from __future__ import annotations

import asyncio
import base64
import json
import mimetypes
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import httpx

from opentulpa.agent.lc_messages import HumanMessage, SystemMessage
from opentulpa.agent.utils import content_to_text as _content_to_text

_VIDEO_EXTENSIONS = {
    ".mp4",
    ".mov",
    ".avi",
    ".webm",
    ".mpeg",
    ".mpg",
    ".m4v",
}
_VIDEO_SEGMENT_SECONDS = 30
_VIDEO_MAX_SEGMENTS = 12
_VIDEO_INLINE_MAX_BYTES = 20 * 1024 * 1024
_VIDEO_FALLBACK_DURATION_SECONDS = 120
_MEDIA_ANALYSIS_RETRIES = 2


async def _ainvoke_runtime_model(runtime: Any, messages: list[Any]) -> Any:
    ainvoke_model = getattr(runtime, "ainvoke_model", None)
    if callable(ainvoke_model):
        return await ainvoke_model(
            runtime._model,
            messages,
            call_context={"call_site": "file_analysis"},
        )
    return await runtime._model.ainvoke(messages)


def _resolve_runtime_model(runtime: Any, *, use_media_model: bool = False) -> tuple[Any, str]:
    if use_media_model:
        media_model = getattr(runtime, "_telegram_media_model", None)
        media_model_name = str(getattr(runtime, "_telegram_media_model_name", "") or "").strip()
        if media_model_name:
            return media_model or getattr(runtime, "_model", None), media_model_name
    model = getattr(runtime, "_model", None)
    model_name = str(getattr(runtime, "model_name", "") or "").strip()
    return model, model_name


async def _ainvoke_selected_runtime_model(
    runtime: Any,
    messages: list[Any],
    *,
    use_media_model: bool = False,
) -> Any:
    model, model_name = _resolve_runtime_model(runtime, use_media_model=use_media_model)
    if model is None:
        raise RuntimeError("runtime model unavailable")
    ainvoke_model = getattr(runtime, "ainvoke_model", None)
    if callable(ainvoke_model):
        return await ainvoke_model(
            model,
            messages,
            model_name=model_name,
            call_context={"call_site": "file_analysis"},
        )
    return await model.ainvoke(messages)


def extract_docx_text(raw_bytes: bytes) -> str:
    try:
        with ZipFile(BytesIO(raw_bytes)) as zf:
            xml_bytes = zf.read("word/document.xml")
    except (BadZipFile, KeyError) as exc:
        raise ValueError("DOCX parsing failed") from exc
    root = ElementTree.fromstring(xml_bytes)
    out: list[str] = []
    for node in root.iter():
        if node.tag.endswith("}t") and node.text:
            out.append(node.text)
    return " ".join(out).strip()


def _spreadsheet_cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def extract_xlsx_text(raw_bytes: bytes, *, max_chars: int = 140000) -> str:
    """Return a bounded, retrieval-friendly workbook preview."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except Exception as exc:
        raise RuntimeError("XLSX parser unavailable. Install openpyxl.") from exc
    try:
        workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"XLSX parsing failed: {exc}") from exc

    parts: list[str] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets[:20], start=1):
            parts.append(
                f"# Sheet {sheet_index}: {sheet.title} "
                f"(rows={sheet.max_row or 0}, columns={sheet.max_column or 0})"
            )
            emitted_rows = 0
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row or 0, 80),
                max_col=min(sheet.max_column or 0, 24),
                values_only=True,
            ):
                cells = [_spreadsheet_cell_to_text(value) for value in row]
                if not any(cells):
                    continue
                while cells and not cells[-1]:
                    cells.pop()
                parts.append(" | ".join(cells))
                emitted_rows += 1
                if len("\n".join(parts)) >= max_chars:
                    break
            if emitted_rows == 0:
                parts.append("(no non-empty rows in preview)")
            parts.append("")
            if len("\n".join(parts)) >= max_chars:
                break
    finally:
        workbook.close()
    return "\n".join(parts).strip()[:max_chars]


def extract_pdf_text(raw_bytes: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError("PDF parser unavailable. Install pypdf.") from exc
    try:
        reader = PdfReader(BytesIO(raw_bytes))
        parts: list[str] = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts).strip()
    except Exception as exc:
        raise ValueError(f"PDF parsing failed: {exc}") from exc


def extract_uploaded_text(
    *,
    raw_bytes: bytes,
    filename: str | None,
    mime_type: str | None,
    max_chars: int = 140000,
) -> str:
    name = str(filename or "").lower()
    mime = str(mime_type or "").lower()
    text = ""
    try:
        if mime.startswith("text/") or any(
            name.endswith(ext)
            for ext in (".txt", ".md", ".csv", ".tsv", ".json", ".yaml", ".yml", ".log")
        ):
            text = raw_bytes.decode("utf-8", errors="replace")
        elif mime == "application/pdf" or name.endswith(".pdf"):
            text = extract_pdf_text(raw_bytes)
        elif (
            mime
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or name.endswith(".docx")
        ):
            text = extract_docx_text(raw_bytes)
        elif (
            mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or name.endswith(".xlsx")
        ):
            text = extract_xlsx_text(raw_bytes, max_chars=max_chars)
    except Exception:
        text = ""
    return str(text or "").strip()[:max_chars]


def _infer_audio_format(*, filename: str | None, mime_type: str | None) -> str:
    safe_name = str(filename or "").lower().strip()
    safe_mime = str(mime_type or "").lower().split(";", 1)[0].strip()
    ext = ""
    if "." in safe_name:
        ext = safe_name.rsplit(".", 1)[-1].strip()

    ext_map = {
        "wav": "wav",
        "mp3": "mp3",
        "aiff": "aiff",
        "aac": "aac",
        "ogg": "ogg",
        "oga": "ogg",
        "flac": "flac",
        "m4a": "m4a",
    }
    if ext in ext_map:
        return ext_map[ext]

    mime_map = {
        "audio/wav": "wav",
        "audio/x-wav": "wav",
        "audio/mpeg": "mp3",
        "audio/mp3": "mp3",
        "audio/aiff": "aiff",
        "audio/aac": "aac",
        "audio/ogg": "ogg",
        "audio/flac": "flac",
        "audio/mp4": "m4a",
        "audio/m4a": "m4a",
    }
    return mime_map.get(safe_mime, "ogg")


def _looks_like_video_blob(*, filename: str, mime_type: str, kind: str) -> bool:
    safe_name = str(filename or "").strip().lower()
    safe_mime = str(mime_type or "").strip().lower()
    safe_kind = str(kind or "").strip().lower()
    if safe_kind in {"video", "video_note"}:
        return True
    if safe_mime.startswith("video/"):
        return True
    return any(safe_name.endswith(ext) for ext in _VIDEO_EXTENSIONS)


def _video_mime_or_default(mime_type: str) -> str:
    safe = str(mime_type or "").strip().lower()
    return safe if safe.startswith("video/") else "video/mp4"


def _seconds_to_mmss(seconds: int) -> str:
    total = max(0, int(seconds))
    mins = total // 60
    secs = total % 60
    return f"{mins:02d}:{secs:02d}"


def _build_30s_segments(*, duration_seconds: int, max_segments: int = _VIDEO_MAX_SEGMENTS) -> list[tuple[int, int]]:
    safe_max = max(1, int(max_segments))
    safe_duration = max(1, int(duration_seconds))
    out: list[tuple[int, int]] = []
    start = 0
    while start < safe_duration and len(out) < safe_max:
        end = min(start + _VIDEO_SEGMENT_SECONDS, safe_duration)
        out.append((start, end))
        start = end
    return out


def _extract_first_json_block(text: str) -> dict[str, Any] | None:
    raw = str(text or "").strip()
    if not raw:
        return None
    for match in re.finditer(r"\{.*?\}", raw, flags=re.DOTALL):
        chunk = match.group(0).strip()
        try:
            payload = json.loads(chunk)
        except Exception:
            continue
        if isinstance(payload, dict):
            return payload
    return None


async def _estimate_video_duration_seconds(
    runtime: Any,
    *,
    video_data_url: str,
    question: str,
    caption: str,
) -> int:
    prompt = (
        "Estimate this video's total duration. "
        "Return JSON only: {\"duration_seconds\": <integer>}."
    )
    if question:
        prompt += f"\nUser question: {question[:400]}"
    if caption:
        prompt += f"\nUser caption: {caption[:400]}"
    try:
        response = await _ainvoke_selected_runtime_model(
            runtime,
            [
                SystemMessage(content="Return strict JSON only."),
                HumanMessage(
                    content=[
                        {"type": "text", "text": prompt},
                        {"type": "video_url", "video_url": {"url": video_data_url}},
                    ]
                ),
            ],
            use_media_model=True,
        )
    except Exception:
        return 0
    text = _content_to_text(getattr(response, "content", "")).strip()
    payload = _extract_first_json_block(text)
    if not payload:
        return 0
    duration = payload.get("duration_seconds")
    try:
        value = int(float(str(duration)))
    except Exception:
        return 0
    return max(0, min(value, 3600))


async def _analyze_video_segment(
    runtime: Any,
    *,
    video_data_url: str,
    start_seconds: int,
    end_seconds: int,
    caption: str,
    question: str,
) -> str:
    start_label = _seconds_to_mmss(start_seconds)
    end_label = _seconds_to_mmss(end_seconds)
    prompt = (
        "Analyze ONLY the requested video window and ignore other timestamps.\n"
        f"Window: {start_label} to {end_label}.\n"
        "Return concise notes with these headings:\n"
        "- Transcript: include any spoken or clearly heard speech in the original language when possible. "
        "Use short quotes or verbatim lines; mark unclear parts as [inaudible].\n"
        "- What happens: concise visual actions and scene changes only.\n"
        "- Audio/background: music, notable sounds, or non-speech audio cues only.\n"
        "- Key visual details: only important appearance/style details that affect understanding."
    )
    if caption:
        prompt += f"\nUser caption: {caption[:500]}"
    if question:
        prompt += f"\nUser question focus: {question[:600]}"
    response = await _ainvoke_selected_runtime_model(
        runtime,
        [
            SystemMessage(content="You are precise about timeline-based video analysis."),
            HumanMessage(
                content=[
                    {"type": "text", "text": prompt},
                    {"type": "video_url", "video_url": {"url": video_data_url}},
                ]
            ),
        ],
        use_media_model=True,
    )
    text = _content_to_text(getattr(response, "content", "")).strip()
    if not text:
        raise ValueError("empty video segment analysis")
    return f"{start_label}-{end_label}\n{text[:1800]}"


async def _analyze_video_segment_with_retries(
    runtime: Any,
    *,
    video_data_url: str,
    start_seconds: int,
    end_seconds: int,
    caption: str,
    question: str,
) -> str:
    last_error = ""
    for attempt in range(_MEDIA_ANALYSIS_RETRIES + 1):
        try:
            return await _analyze_video_segment(
                runtime,
                video_data_url=video_data_url,
                start_seconds=start_seconds,
                end_seconds=end_seconds,
                caption=caption,
                question=question,
            )
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt < _MEDIA_ANALYSIS_RETRIES:
                await asyncio.sleep(0.4 * (attempt + 1))
    start_label = _seconds_to_mmss(start_seconds)
    end_label = _seconds_to_mmss(end_seconds)
    return f"{start_label}-{end_label}: segment analysis failed after retries ({last_error})."


async def _synthesize_video_segments(
    runtime: Any,
    *,
    filename: str,
    mime_type: str,
    caption: str,
    question: str,
    segment_notes: list[str],
) -> str:
    compiled_notes = "\n\n".join(note for note in segment_notes if str(note).strip())
    if not compiled_notes:
        return ""
    prompt = (
        "Create a final video report from segmented notes.\n"
        "Output sections:\n"
        "1) Transcript\n"
        "Include the spoken/heard speech in order. Preserve the original language when possible. "
        "Use [inaudible] for unclear fragments.\n"
        "2) What happens\n"
        "A concise description of the visible actions and scene changes.\n"
        "3) Audio/background\n"
        "Only notable non-speech sounds, music, or ambient cues.\n"
        "4) Short overall summary\n"
        "Keep non-transcript description concise and retrieval-friendly."
    )
    if question:
        prompt += f"\nUser question to prioritize: {question[:800]}"
    response = await _ainvoke_selected_runtime_model(
        runtime,
        [
            SystemMessage(content="Synthesize segment notes into one cohesive video report."),
            HumanMessage(
                content=(
                    f"{prompt}\n\n"
                    f"filename={filename}\n"
                    f"mime_type={mime_type}\n"
                    f"caption={caption[:500]}\n\n"
                    "Segment notes:\n"
                    f"{compiled_notes[:24000]}"
                )
            ),
        ],
        use_media_model=True,
    )
    final_text = _content_to_text(getattr(response, "content", "")).strip()
    return final_text[:6000]


async def _summarize_video_blob(
    runtime: Any,
    *,
    filename: str,
    mime_type: str,
    raw_bytes: bytes,
    caption: str,
    question: str,
) -> str:
    content_bytes = bytes(raw_bytes or b"")
    if len(content_bytes) > _VIDEO_INLINE_MAX_BYTES:
        return (
            f"Uploaded video '{filename}' is {len(content_bytes)} bytes, which is too large for inline "
            "video analysis. Please send a shorter/compressed clip (under 20MiB) or share a supported video URL."
        )

    safe_mime = _video_mime_or_default(mime_type)
    b64_video = base64.b64encode(content_bytes).decode("ascii")
    video_data_url = f"data:{safe_mime};base64,{b64_video}"

    estimated_duration = await _estimate_video_duration_seconds(
        runtime,
        video_data_url=video_data_url,
        question=question,
        caption=caption,
    )
    if estimated_duration <= 0:
        estimated_duration = _VIDEO_FALLBACK_DURATION_SECONDS

    segments = _build_30s_segments(duration_seconds=estimated_duration, max_segments=_VIDEO_MAX_SEGMENTS)
    tasks = [
        _analyze_video_segment_with_retries(
            runtime,
            video_data_url=video_data_url,
            start_seconds=start,
            end_seconds=end,
            caption=caption,
            question=question,
        )
        for start, end in segments
    ]
    try:
        raw_notes = await asyncio.gather(*tasks, return_exceptions=True)
    except Exception:
        raw_notes = []
    notes: list[str] = []
    for item in raw_notes:
        if isinstance(item, Exception):
            continue
        text = str(item or "").strip()
        if text:
            notes.append(text)
    return await _synthesize_video_segments(
        runtime,
        filename=filename,
        mime_type=safe_mime,
        caption=caption,
        question=question,
        segment_notes=notes,
    )


async def transcribe_audio_blob(
    runtime: Any,
    *,
    filename: str | None,
    mime_type: str | None,
    kind: str | None,
    raw_bytes: bytes,
) -> str:
    """Transcribe short uploaded audio/voice files via OpenRouter input_audio."""
    safe_kind = str(kind or "").strip().lower()
    safe_mime = str(mime_type or "").lower().split(";", 1)[0].strip()
    content_bytes = bytes(raw_bytes or b"")
    if not content_bytes:
        return ""
    if safe_kind not in {"voice", "audio"} and not safe_mime.startswith("audio/"):
        return ""
    if len(content_bytes) > 12_000_000:
        return ""

    api_key = str(getattr(runtime, "openrouter_api_key", "") or "").strip()
    if not api_key:
        return ""
    base_url = (
        str(getattr(runtime, "openrouter_base_url", "") or "").strip().rstrip("/")
        or "https://openrouter.ai/api/v1"
    )
    _, model_name = _resolve_runtime_model(runtime, use_media_model=True)
    if not model_name:
        return ""

    audio_format = _infer_audio_format(filename=filename, mime_type=safe_mime)
    b64_audio = base64.b64encode(content_bytes).decode("ascii")
    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "Transcribe all spoken or clearly heard speech in this audio accurately. "
                            "Preserve the original language when possible. "
                            "Use [inaudible] for unclear fragments. "
                            "Return plain text transcript only, no commentary or summary."
                        ),
                    },
                    {
                        "type": "input_audio",
                        "input_audio": {
                            "data": b64_audio,
                            "format": audio_format,
                        },
                    },
                ],
            }
        ],
        "temperature": 0,
    }

    try:
        async with httpx.AsyncClient(timeout=75.0) as client:
            resp = await client.post(
                f"{base_url}/chat/completions",
                json=payload,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
            )
        if resp.status_code >= 400:
            return ""
        data = resp.json()
        choices = data.get("choices", [])
        if not isinstance(choices, list) or not choices:
            return ""
        message = choices[0].get("message", {}) if isinstance(choices[0], dict) else {}
        transcript = _content_to_text(message.get("content", "")).strip()
        return transcript[:4000]
    except Exception:
        return ""


async def summarize_uploaded_blob(
    runtime: Any,
    *,
    filename: str | None,
    mime_type: str | None,
    kind: str | None,
    raw_bytes: bytes,
    caption: str | None = None,
    question: str | None = None,
) -> str:
    safe_filename = str(filename or "file.bin").strip() or "file.bin"
    safe_mime = str(mime_type or "").strip().lower()
    if not safe_mime:
        guessed, _ = mimetypes.guess_type(safe_filename)
        safe_mime = str(guessed or "").strip().lower()
    safe_kind = str(kind or "file").strip() or "file"
    q = str(question or "").strip()
    caption_text = str(caption or "").strip()
    content_bytes = bytes(raw_bytes or b"")
    if not content_bytes:
        return f"Uploaded {safe_kind} file '{safe_filename}' was empty."

    if _looks_like_video_blob(filename=safe_filename, mime_type=safe_mime, kind=safe_kind):
        with_video = await _summarize_video_blob(
            runtime,
            filename=safe_filename,
            mime_type=safe_mime,
            raw_bytes=content_bytes,
            caption=caption_text,
            question=q,
        )
        if with_video:
            return with_video[:6000]

    if safe_kind in {"voice", "audio"} or safe_mime.startswith("audio/"):
        transcript = await transcribe_audio_blob(
            runtime,
            filename=safe_filename,
            mime_type=safe_mime or None,
            kind=safe_kind,
            raw_bytes=content_bytes,
        )
        if transcript:
            return (
                f"Uploaded {safe_kind or 'audio'} file '{safe_filename}'. "
                f"Transcript: {transcript[:5000]}"
            )

    # Gemini/OpenRouter can handle image input; keep payload bounded to avoid excessive prompt size.
    if safe_mime.startswith("image/") and len(content_bytes) <= 2_000_000:
        try:
            b64 = base64.b64encode(content_bytes).decode("ascii")
            data_url = f"data:{safe_mime};base64,{b64}"
            prompt_text = (
                "Analyze this uploaded image and summarize key information. "
                "Extract visible text, tables, IDs, dates, totals, names, and action items if present. "
                "Keep the summary concise and retrieval-friendly."
            )
            if q:
                prompt_text += f"\nUser question about this file: {q}"
            if caption_text:
                prompt_text += f"\nUser caption: {caption_text[:500]}"
            response = await _ainvoke_selected_runtime_model(
                runtime,
                [
                    SystemMessage(content="You analyze uploaded user files accurately."),
                    HumanMessage(
                        content=[
                            {"type": "text", "text": prompt_text},
                            {"type": "image_url", "image_url": {"url": data_url}},
                        ]
                    ),
                ],
                use_media_model=True,
            )
            vision_summary = _content_to_text(getattr(response, "content", "")).strip()
            if vision_summary:
                return vision_summary[:6000]
        except Exception:
            pass

    extracted = extract_uploaded_text(
        raw_bytes=content_bytes,
        filename=safe_filename,
        mime_type=safe_mime,
        max_chars=140000,
    )
    if extracted:
        prompt = (
            "Summarize this uploaded file for future retrieval. Include key facts, entities, "
            "dates, amounts, and concise keywords."
        )
        if q:
            prompt = (
                "Answer the user's question using only this uploaded file content. "
                "If uncertain, say what is missing."
            )
        response = await _ainvoke_runtime_model(
            runtime,
            [
                SystemMessage(content="You analyze uploaded file content accurately and concisely."),
                HumanMessage(
                    content=(
                        f"filename={safe_filename}\n"
                        f"mime_type={safe_mime or 'unknown'}\n"
                        f"kind={safe_kind}\n"
                        f"caption={caption_text[:500]}\n"
                        f"question={q[:800]}\n\n"
                        f"{prompt}\n\n"
                        "File content:\n"
                        f"{extracted}"
                    )
                ),
            ],
        )
        text_summary = _content_to_text(getattr(response, "content", "")).strip()
        if text_summary:
            return text_summary[:6000]

    return (
        f"Uploaded {safe_kind} file '{safe_filename}' "
        f"(mime={safe_mime or 'unknown'}, size_bytes={len(content_bytes)}). "
        "No extractable text was available."
    )


async def analyze_uploaded_file(
    runtime: Any,
    *,
    record: dict[str, Any],
    raw_bytes: bytes,
    question: str | None = None,
) -> dict[str, Any]:
    analysis = await summarize_uploaded_blob(
        runtime,
        filename=str(record.get("original_filename", "")).strip() or None,
        mime_type=str(record.get("mime_type", "")).strip() or None,
        kind=str(record.get("kind", "")).strip() or None,
        raw_bytes=raw_bytes,
        caption=str(record.get("caption", "")).strip() or None,
        question=question,
    )
    return {
        "file_id": str(record.get("id", "")).strip(),
        "analysis": str(analysis or "").strip()[:6000],
        "question": str(question or "").strip() or None,
    }
