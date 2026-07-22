"""Inspect uploaded source files so workflow setup can choose useful inputs."""

from __future__ import annotations

import csv
import re
from collections.abc import Iterable
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from opentulpa.agent.file_analysis import extract_uploaded_text

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MAX_CELL_CHARS = 300
_MAX_STRUCTURE_SHEETS = 40
_MAX_STRUCTURE_SAMPLE_ROWS = 8
_MAX_STRUCTURE_SAMPLE_COLS = 12
_MAX_MATCHES_PER_SHEET = 12
_MAX_TABLE_CANDIDATES_PER_SHEET = 12


def normalize_hints(value: Any) -> list[str]:
    if isinstance(value, str):
        candidates: Iterable[Any] = re.split(r"[\n,;]+", value)
    elif isinstance(value, list):
        candidates = value
    else:
        candidates = []
    out: list[str] = []
    seen: set[str] = set()
    for item in candidates:
        text = str(item or "").strip()
        folded = text.casefold()
        if not text or folded in seen:
            continue
        seen.add(folded)
        out.append(text[:120])
    return out[:20]


def inspect_uploaded_file_structure(
    *,
    raw_bytes: bytes,
    filename: str | None,
    mime_type: str | None,
    search_terms: Any = None,
) -> dict[str, Any]:
    """Return a compact structural map so an agent can choose useful regions."""
    terms = normalize_hints(search_terms)
    safe_filename = str(filename or "file.bin").strip() or "file.bin"
    safe_mime = str(mime_type or "").strip().lower()
    name = safe_filename.lower()
    if safe_mime == _XLSX_MIME or name.endswith(".xlsx"):
        return _inspect_xlsx_structure(
            raw_bytes=raw_bytes,
            filename=safe_filename,
            search_terms=terms,
        )
    if name.endswith(".csv") or safe_mime in {"text/csv", "application/csv"}:
        return _inspect_delimited_text(raw_bytes=raw_bytes, filename=safe_filename, format_name="csv")

    extracted = extract_uploaded_text(
        raw_bytes=raw_bytes,
        filename=safe_filename,
        mime_type=safe_mime,
        max_chars=16_000,
    )
    return {
        "filename": safe_filename,
        "mime_type": safe_mime,
        "format": "text" if extracted else "unknown",
        "warnings": [] if extracted else [f"no extractable text for {safe_filename}"],
        "structure": {
            "line_count": len(str(extracted or "").splitlines()),
            "preview": str(extracted or "")[:4000],
        },
    }


def _inspect_xlsx_structure(
    *,
    raw_bytes: bytes,
    filename: str,
    search_terms: list[str],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except Exception as exc:  # pragma: no cover - exercised only without dependency
        return {
            "filename": filename,
            "mime_type": _XLSX_MIME,
            "format": "xlsx",
            "warnings": [f"xlsx parser unavailable for {filename}: {exc}"],
            "structure": {"sheets": []},
        }

    warnings: list[str] = []
    try:
        workbook = load_workbook(BytesIO(raw_bytes), data_only=True, read_only=False)
    except Exception as exc:
        return {
            "filename": filename,
            "mime_type": _XLSX_MIME,
            "format": "xlsx",
            "warnings": [f"xlsx parsing failed for {filename}: {exc}"],
            "structure": {"sheets": []},
        }

    sheets: list[dict[str, Any]] = []
    sheet_names = list(workbook.sheetnames)
    if len(sheet_names) > _MAX_STRUCTURE_SHEETS:
        warnings.append(f"sheet inventory truncated to {_MAX_STRUCTURE_SHEETS}")
    for index, sheet_name in enumerate(sheet_names[:_MAX_STRUCTURE_SHEETS], start=1):
        sheet = workbook[sheet_name]
        merged_values = _merged_cell_values(sheet)
        nonempty_rows: list[tuple[int, list[str]]] = []
        matches: list[dict[str, Any]] = []
        matched_terms = [
            term for term in search_terms if _matches_any_hint(sheet_name, [term])
        ]
        for row_index in range(1, int(sheet.max_row or 0) + 1):
            values = [
                _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
                for col_index in range(1, min(int(sheet.max_column or 0), _MAX_STRUCTURE_SAMPLE_COLS) + 1)
            ]
            full_values = values
            if int(sheet.max_column or 0) > _MAX_STRUCTURE_SAMPLE_COLS and search_terms:
                full_values = [
                    _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
                    for col_index in range(1, int(sheet.max_column or 0) + 1)
                ]
            if any(values) or any(full_values):
                nonempty_rows.append((row_index, values))
            if search_terms and len(matches) < _MAX_MATCHES_PER_SHEET:
                row_text = " ".join(full_values)
                matched_term = next((term for term in search_terms if _matches_any_hint(row_text, [term])), "")
                if matched_term:
                    matches.append(
                        {
                            "term": matched_term,
                            "source_ref": f"{sheet_name}!{row_index}",
                            "row": row_index,
                            "values": _trim_row_values(full_values),
                        }
                    )
        include_details = not search_terms or bool(matched_terms)
        sample_rows = [
            {"source_ref": f"{sheet_name}!{row}", "row": row, "values": _trim_row_values(values)}
            for row, values in nonempty_rows[:_MAX_STRUCTURE_SAMPLE_ROWS]
        ] if include_details else []
        table_candidates = _table_candidates(sheet_name, nonempty_rows) if include_details else []
        visible_matches = matches if include_details else matches[:3]
        omitted_reason = ""
        if not include_details:
            omitted_reason = (
                "sample rows and table candidates omitted because this sheet did not match "
                "search_terms by sheet name; see matches for relevant cell hits"
            )
        sheets.append(
            {
                "index": index,
                "name": sheet_name,
                "matched_terms": matched_terms,
                "max_row": int(sheet.max_row or 0),
                "max_column": int(sheet.max_column or 0),
                "nonempty_rows": len(nonempty_rows),
                "sample_rows": sample_rows,
                "table_candidates": table_candidates,
                "matches": visible_matches,
                **({"omitted_detail_reason": omitted_reason} if omitted_reason else {}),
            }
        )

    return {
        "filename": filename,
        "mime_type": _XLSX_MIME,
        "format": "xlsx",
        "warnings": warnings,
        "structure": {
            "sheets": sheets,
            "selection_format": {
                "file_id": "optional source file id when preparing multiple files",
                "sheet_name": "exact sheet name from this inventory",
                "row_start": "optional 1-based first row",
                "row_end": "optional 1-based last row",
            },
        },
    }


def _inspect_delimited_text(
    *,
    raw_bytes: bytes,
    filename: str,
    format_name: str,
) -> dict[str, Any]:
    text = raw_bytes.decode("utf-8", errors="replace")
    rows = list(csv.reader(StringIO(text)))[:_MAX_STRUCTURE_SAMPLE_ROWS]
    return {
        "filename": filename,
        "mime_type": "text/csv",
        "format": format_name,
        "warnings": [],
        "structure": {
            "sample_rows": [
                {"row": index, "values": _trim_row_values(row)}
                for index, row in enumerate(rows, start=1)
            ],
        },
    }


def _table_candidates(sheet_name: str, nonempty_rows: list[tuple[int, list[str]]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    current: list[tuple[int, list[str]]] = []
    previous_row = 0
    for row_number, values in nonempty_rows:
        if current and row_number > previous_row + 1:
            candidates.append(_table_candidate(sheet_name, current))
            current = []
        current.append((row_number, values))
        previous_row = row_number
    if current:
        candidates.append(_table_candidate(sheet_name, current))
    return candidates[:_MAX_TABLE_CANDIDATES_PER_SHEET]


def _table_candidate(sheet_name: str, rows: list[tuple[int, list[str]]]) -> dict[str, Any]:
    first = rows[0][0]
    last = rows[-1][0]
    sample = rows[: min(len(rows), 3)]
    return {
        "sheet_name": sheet_name,
        "row_start": first,
        "row_end": last,
        "nonempty_rows": len(rows),
        "sample_rows": [
            {"source_ref": f"{sheet_name}!{row}", "row": row, "values": _trim_row_values(values)}
            for row, values in sample
        ],
    }


def _merged_cell_values(sheet: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    ranges = getattr(getattr(sheet, "merged_cells", None), "ranges", []) or []
    for merged in ranges:
        top_value = sheet.cell(int(merged.min_row), int(merged.min_col)).value
        if top_value is None:
            continue
        for row_index in range(int(merged.min_row), int(merged.max_row) + 1):
            for col_index in range(int(merged.min_col), int(merged.max_col) + 1):
                values[(row_index, col_index)] = top_value
    return values


def _cell_value(sheet: Any, row_index: int, col_index: int, merged_values: dict[tuple[int, int], Any]) -> Any:
    value = sheet.cell(row_index, col_index).value
    if value is None:
        return merged_values.get((row_index, col_index))
    return value


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_inline(value, limit=_MAX_CELL_CHARS)


def _clean_inline(value: Any, *, limit: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text[:limit]


def _matches_any_hint(value: Any, hints: list[str]) -> bool:
    if not hints:
        return False
    text = str(value or "").casefold()
    return any(hint.casefold() in text or text in hint.casefold() for hint in hints if hint)


def _trim_row_values(values: list[Any]) -> list[str]:
    return [
        _clean_inline(value, limit=_MAX_CELL_CHARS)
        for value in values[:_MAX_STRUCTURE_SAMPLE_COLS]
    ]
