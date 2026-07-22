"""Source-file normalization for workflow business knowledge."""

from __future__ import annotations

import csv
import json
import re
from collections.abc import Iterable
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO, StringIO
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from kobo.business_knowledge.models import KnowledgeSourceSection

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOCX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

_TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".json",
    ".yaml",
    ".yml",
    ".log",
    ".html",
    ".htm",
}
_CSV_EXTENSIONS = {".csv", ".tsv"}
_MAX_TEXT_SECTION_CHARS = 8000
_MAX_TABLE_SECTION_CHARS = 16000
_MAX_CELL_CHARS = 500
_MAX_ROWS_PER_SHEET = 20000
_MAX_COLS_PER_SHEET = 200
_HEADER_CONTEXT_ROW_LIMIT = 12


def content_hash(raw_bytes: bytes) -> str:
    return sha256(bytes(raw_bytes or b"")).hexdigest()


def metadata_json(metadata: dict[str, Any]) -> str:
    return json.dumps(metadata, ensure_ascii=False, sort_keys=True)


def extract_source_sections(
    *,
    record: dict[str, Any],
    raw_bytes: bytes,
) -> tuple[list[KnowledgeSourceSection], list[str], str]:
    """Normalize one uploaded file into source-pack sections."""

    filename = str(record.get("original_filename", "") or "file.bin").strip() or "file.bin"
    mime_type = str(record.get("mime_type", "") or "").strip().lower()
    lower_name = filename.lower()
    file_id = str(record.get("id", "") or "").strip()

    if mime_type == XLSX_MIME_TYPE or lower_name.endswith(".xlsx"):
        return _extract_xlsx_sections(file_id=file_id, filename=filename, raw_bytes=raw_bytes)
    if _is_csv_like(lower_name=lower_name, mime_type=mime_type):
        return _extract_delimited_sections(file_id=file_id, filename=filename, raw_bytes=raw_bytes)
    if mime_type == "application/pdf" or lower_name.endswith(".pdf"):
        sections, warnings, source_kind = _extract_pdf_sections(
            file_id=file_id,
            filename=filename,
            raw_bytes=raw_bytes,
        )
        derived = _derived_media_text(record)
        if sections and derived:
            derived_sections, derived_warnings, _derived_source_kind = _derived_media_sections(
                record=record,
                file_id=file_id,
                filename=filename,
                mime_type=mime_type,
                base_sort_order=len(sections),
            )
            return sections + derived_sections, warnings + derived_warnings, source_kind
        if sections or not derived:
            return sections, warnings, source_kind
        derived_sections, derived_warnings, derived_source_kind = _derived_media_sections(
            record=record,
            file_id=file_id,
            filename=filename,
            mime_type=mime_type,
        )
        return derived_sections, warnings + derived_warnings, derived_source_kind
    if mime_type == DOCX_MIME_TYPE or lower_name.endswith(".docx"):
        return _extract_docx_sections(file_id=file_id, filename=filename, raw_bytes=raw_bytes)
    if mime_type.startswith("text/") or any(lower_name.endswith(ext) for ext in _TEXT_EXTENSIONS):
        text = raw_bytes.decode("utf-8", errors="replace")
        sections = _text_sections(
            file_id=file_id,
            filename=filename,
            text=text,
            metadata={
                "document_title": filename,
                "section_title": filename,
                "heading_path": [filename],
                "format": "markdown" if lower_name.endswith((".md", ".markdown")) else "text",
                "mime_type": mime_type or "text/plain",
            },
        )
        return sections, [], "local_source"

    if _derived_media_text(record):
        return _derived_media_sections(record=record, file_id=file_id, filename=filename, mime_type=mime_type)

    return [], [f"unsupported business-knowledge file type: {filename}"], "unsupported_for_business_knowledge"


def _is_csv_like(*, lower_name: str, mime_type: str) -> bool:
    return (
        lower_name.endswith(tuple(_CSV_EXTENSIONS))
        or mime_type in {"text/csv", "application/csv", "text/tab-separated-values"}
    )


def _derived_media_text(record: dict[str, Any]) -> str:
    summary = str(record.get("summary", "") or "").strip()
    if "ai_summary=" in summary:
        return summary.split("ai_summary=", 1)[1].strip()[:6000]
    text_excerpt = str(record.get("text_excerpt", "") or "").strip()
    return text_excerpt[:6000] if text_excerpt else ""


def _derived_media_sections(
    *,
    record: dict[str, Any],
    file_id: str,
    filename: str,
    mime_type: str,
    base_sort_order: int = 0,
) -> tuple[list[KnowledgeSourceSection], list[str], str]:
    derived = _derived_media_text(record)
    return (
        [
            KnowledgeSourceSection(
                content=(
                    f"Document: {filename}\n"
                    "Derived media analysis. Treat this as non-authoritative for exact "
                    "prices, policies, and service menus unless confirmed elsewhere.\n"
                    f"{derived}"
                ),
                source_ref=f"{file_id}:derived_media_summary",
                source_kind="derived_from_media",
                sort_order=base_sort_order + 1,
                metadata={
                    "file_id": file_id,
                    "filename": filename,
                    "document_title": filename,
                    "section_title": filename,
                    "heading_path": [filename],
                    "source_label": f"{filename} derived media summary",
                    "locator": "derived media summary",
                    "mime_type": mime_type,
                    "derived": True,
                },
            )
        ],
        [
            "prepared existing media analysis as derived evidence; exact prices, policies, and service menus need owner confirmation"
        ],
        "derived_from_media",
    )


def _extract_xlsx_sections(
    *,
    file_id: str,
    filename: str,
    raw_bytes: bytes,
) -> tuple[list[KnowledgeSourceSection], list[str], str]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
        from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
    except Exception as exc:
        return [], [f"xlsx parser unavailable for {filename}: {exc}"], "unsupported_for_business_knowledge"

    warnings: list[str] = []
    try:
        workbook = load_workbook(BytesIO(raw_bytes), data_only=True, read_only=False)
    except Exception as exc:
        return [], [f"xlsx parsing failed for {filename}: {exc}"], "unsupported_for_business_knowledge"

    sections: list[KnowledgeSourceSection] = []
    sort_order = 0
    for sheet in workbook.worksheets:
        sheet_name = str(sheet.title or "").strip() or "Sheet"
        merged_values = _merged_cell_values(sheet)
        nonempty_rows: list[tuple[int, list[str]]] = []
        max_row = min(int(sheet.max_row or 0), _MAX_ROWS_PER_SHEET)
        max_col = min(int(sheet.max_column or 0), _MAX_COLS_PER_SHEET)
        if int(sheet.max_row or 0) > _MAX_ROWS_PER_SHEET:
            warnings.append(f"{sheet_name}: row count truncated at {_MAX_ROWS_PER_SHEET} for knowledge pack")
        if int(sheet.max_column or 0) > _MAX_COLS_PER_SHEET:
            warnings.append(f"{sheet_name}: column count truncated at {_MAX_COLS_PER_SHEET} for knowledge pack")
        for row_index in range(1, max_row + 1):
            values = [
                _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
                for col_index in range(1, max_col + 1)
            ]
            if any(values):
                nonempty_rows.append((row_index, values))
        if not nonempty_rows:
            continue
        header_context = _header_context_rows(nonempty_rows)
        current_lines: list[str] = []
        current_start = nonempty_rows[0][0]
        current_end = current_start
        for row_index, values in nonempty_rows:
            line = _xlsx_row_line(row_index=row_index, values=values, get_column_letter=get_column_letter)
            if (
                current_lines
                and sum(len(item) + 1 for item in current_lines) + len(line) + 1 > _MAX_TABLE_SECTION_CHARS
            ):
                sort_order += 1
                sections.append(
                    _table_section(
                        file_id=file_id,
                        filename=filename,
                        sheet_name=sheet_name,
                        header_context=header_context,
                        row_lines=current_lines,
                        row_start=current_start,
                        row_end=current_end,
                        sort_order=sort_order,
                    )
                )
                current_lines = []
                current_start = row_index
            current_lines.append(line)
            current_end = row_index
        if current_lines:
            sort_order += 1
            sections.append(
                _table_section(
                    file_id=file_id,
                    filename=filename,
                    sheet_name=sheet_name,
                    header_context=header_context,
                    row_lines=current_lines,
                    row_start=current_start,
                    row_end=current_end,
                    sort_order=sort_order,
                )
            )
    return sections, warnings, "structured_table"


def _table_section(
    *,
    file_id: str,
    filename: str,
    sheet_name: str,
    header_context: list[str],
    row_lines: list[str],
    row_start: int,
    row_end: int,
    sort_order: int,
) -> KnowledgeSourceSection:
    source_ref = f"{file_id}:{sheet_name}:rows:{row_start}-{row_end}"
    content = (
        f"Workbook: {filename}\n"
        f"Sheet: {sheet_name}\n"
        f"Rows: {row_start}-{row_end}\n"
        "Header/table context rows:\n"
        + "\n".join(header_context)
        + "\n\nRows in this section:\n"
        + "\n".join(row_lines)
    ).strip()
    return KnowledgeSourceSection(
        content=content,
        source_ref=source_ref,
        source_kind="structured_table",
        sort_order=sort_order,
        metadata={
            "file_id": file_id,
            "filename": filename,
            "document_title": filename,
            "format": "xlsx",
            "section_title": sheet_name,
            "heading_path": [sheet_name],
            "source_label": f"{sheet_name} rows {row_start}-{row_end}",
            "locator": f"rows {row_start}-{row_end}",
            "sheet": sheet_name,
            "row_start": row_start,
            "row_end": row_end,
        },
    )


def _xlsx_row_line(*, row_index: int, values: list[str], get_column_letter: Any) -> str:
    parts: list[str] = []
    for col_index, value in enumerate(values, start=1):
        if not value:
            continue
        parts.append(f"{get_column_letter(col_index)}={value}")
    return f"Row {row_index}: " + " | ".join(parts)


def _header_context_rows(rows: list[tuple[int, list[str]]]) -> list[str]:
    candidates: list[tuple[int, list[str]]] = []
    for row_index, values in rows[:_HEADER_CONTEXT_ROW_LIMIT]:
        filled = [value for value in values if value]
        if len(filled) >= 2 or (filled and row_index <= 3):
            candidates.append((row_index, values))
    if not candidates:
        candidates = rows[: min(3, len(rows))]
    out: list[str] = []
    for row_index, values in candidates[:_HEADER_CONTEXT_ROW_LIMIT]:
        parts = [f"col_{idx}={value}" for idx, value in enumerate(values, start=1) if value]
        if parts:
            out.append(f"Row {row_index}: " + " | ".join(parts))
    return out


def _extract_delimited_sections(
    *,
    file_id: str,
    filename: str,
    raw_bytes: bytes,
) -> tuple[list[KnowledgeSourceSection], list[str], str]:
    text = raw_bytes.decode("utf-8", errors="replace")
    delimiter = "\t" if filename.lower().endswith(".tsv") else ","
    rows = list(csv.reader(StringIO(text), delimiter=delimiter))
    nonempty: list[tuple[int, list[str]]] = []
    for row_index, row in enumerate(rows, start=1):
        values = [str(cell or "").strip()[:_MAX_CELL_CHARS] for cell in row]
        if any(values):
            nonempty.append((row_index, values))
    if not nonempty:
        return [], [], "structured_table"
    header_context = _header_context_rows(nonempty)
    sections: list[KnowledgeSourceSection] = []
    current: list[str] = []
    current_start = nonempty[0][0]
    current_end = current_start
    sort_order = 0
    for row_index, values in nonempty:
        line = f"Row {row_index}: " + " | ".join(
            f"col_{col_index}={value}"
            for col_index, value in enumerate(values, start=1)
            if value
        )
        if current and sum(len(item) + 1 for item in current) + len(line) + 1 > _MAX_TABLE_SECTION_CHARS:
            sort_order += 1
            sections.append(
                _delimited_section(
                    file_id=file_id,
                    filename=filename,
                    header_context=header_context,
                    row_lines=current,
                    row_start=current_start,
                    row_end=current_end,
                    format_name="csv" if delimiter == "," else "tsv",
                    sort_order=sort_order,
                )
            )
            current = []
            current_start = row_index
        current.append(line)
        current_end = row_index
    if current:
        sort_order += 1
        sections.append(
            _delimited_section(
                file_id=file_id,
                filename=filename,
                header_context=header_context,
                row_lines=current,
                row_start=current_start,
                row_end=current_end,
                format_name="csv" if delimiter == "," else "tsv",
                sort_order=sort_order,
            )
        )
    return sections, [], "structured_table"


def _delimited_section(
    *,
    file_id: str,
    filename: str,
    header_context: list[str],
    row_lines: list[str],
    row_start: int,
    row_end: int,
    format_name: str,
    sort_order: int,
) -> KnowledgeSourceSection:
    source_ref = f"{file_id}:rows:{row_start}-{row_end}"
    content = (
        f"Table: {filename}\n"
        f"Rows: {row_start}-{row_end}\n"
        "Header/table context rows:\n"
        + "\n".join(header_context)
        + "\n\nRows in this section:\n"
        + "\n".join(row_lines)
    ).strip()
    return KnowledgeSourceSection(
        content=content,
        source_ref=source_ref,
        source_kind="structured_table",
        sort_order=sort_order,
        metadata={
            "file_id": file_id,
            "filename": filename,
            "document_title": filename,
            "section_title": filename,
            "heading_path": [filename],
            "source_label": f"{filename} rows {row_start}-{row_end}",
            "locator": f"rows {row_start}-{row_end}",
            "format": format_name,
            "row_start": row_start,
            "row_end": row_end,
        },
    )


def _extract_pdf_sections(
    *,
    file_id: str,
    filename: str,
    raw_bytes: bytes,
) -> tuple[list[KnowledgeSourceSection], list[str], str]:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        return [], [f"pdf parser unavailable for {filename}: {exc}"], "unsupported_for_business_knowledge"
    try:
        reader = PdfReader(BytesIO(raw_bytes))
    except Exception as exc:
        return [], [f"pdf parsing failed for {filename}: {exc}"], "unsupported_for_business_knowledge"
    sections: list[KnowledgeSourceSection] = []
    sort_order = 0
    for page_index, page in enumerate(reader.pages, start=1):
        text = str(page.extract_text() or "").strip()
        if not text:
            continue
        for part_index, text_part in enumerate(_split_text(text, max_chars=_MAX_TEXT_SECTION_CHARS), start=1):
            sort_order += 1
            sections.append(
                KnowledgeSourceSection(
                    content=f"Document: {filename}\nPage: {page_index}\n{text_part}",
                    source_ref=f"{file_id}:page:{page_index}:part:{part_index}",
                    source_kind="local_source",
                    sort_order=sort_order,
                    metadata={
                        "file_id": file_id,
                        "filename": filename,
                        "document_title": filename,
                        "section_title": filename,
                        "heading_path": [filename],
                        "source_label": f"{filename} page {page_index}",
                        "locator": f"page {page_index}",
                        "format": "pdf",
                        "page": page_index,
                    },
                )
            )
    return sections, [], "local_source"


def _extract_docx_sections(
    *,
    file_id: str,
    filename: str,
    raw_bytes: bytes,
) -> tuple[list[KnowledgeSourceSection], list[str], str]:
    try:
        with ZipFile(BytesIO(raw_bytes)) as zf:
            xml_bytes = zf.read("word/document.xml")
    except (BadZipFile, KeyError) as exc:
        return [], [f"docx parsing failed for {filename}: {exc}"], "unsupported_for_business_knowledge"
    try:
        root = ElementTree.fromstring(xml_bytes)
    except Exception as exc:
        return [], [f"docx XML parsing failed for {filename}: {exc}"], "unsupported_for_business_knowledge"
    paragraphs: list[str] = []
    for node in root.iter():
        if node.tag.endswith("}t") and node.text:
            paragraphs.append(str(node.text).strip())
    text = " ".join(part for part in paragraphs if part).strip()
    sections = _text_sections(
        file_id=file_id,
        filename=filename,
        text=text,
        metadata={
            "document_title": filename,
            "section_title": filename,
            "heading_path": [filename],
            "format": "docx",
        },
    )
    return sections, [], "local_source"


def _text_sections(
    *,
    file_id: str,
    filename: str,
    text: str,
    metadata: dict[str, Any],
) -> list[KnowledgeSourceSection]:
    sections: list[KnowledgeSourceSection] = []
    lines = str(text or "").splitlines()
    current: list[str] = []
    start_line = 1
    current_len = 0
    sort_order = 0
    for line_number, line in enumerate(lines, start=1):
        clean = str(line or "").rstrip()
        if not clean and not current:
            continue
        if current and current_len + len(clean) + 1 > _MAX_TEXT_SECTION_CHARS:
            sort_order += 1
            sections.append(
                _text_section(
                    file_id,
                    filename,
                    current,
                    start_line,
                    line_number - 1,
                    _metadata_for_text_range(metadata, lines, start_line, line_number - 1),
                    sort_order,
                )
            )
            current = []
            current_len = 0
            start_line = line_number
        current.append(clean)
        current_len += len(clean) + 1
    if current:
        sort_order += 1
        sections.append(
            _text_section(
                file_id,
                filename,
                current,
                start_line,
                len(lines),
                _metadata_for_text_range(metadata, lines, start_line, len(lines)),
                sort_order,
            )
        )
    return sections


def _text_section(
    file_id: str,
    filename: str,
    lines: list[str],
    start_line: int,
    end_line: int,
    metadata: dict[str, Any],
    sort_order: int,
) -> KnowledgeSourceSection:
    text = "\n".join(lines).strip()
    source_ref = f"{file_id}:lines:{start_line}-{end_line}"
    return KnowledgeSourceSection(
        content=f"Document: {filename}\nLines: {start_line}-{end_line}\n{text}",
        source_ref=source_ref,
        source_kind="local_source",
        sort_order=sort_order,
        metadata={
            "file_id": file_id,
            "filename": filename,
            "source_label": f"{filename} lines {start_line}-{end_line}",
            "locator": f"lines {start_line}-{end_line}",
            "line_start": start_line,
            "line_end": end_line,
            **metadata,
        },
    )


def _metadata_for_text_range(
    metadata: dict[str, Any],
    all_lines: list[str],
    start_line: int,
    end_line: int,
) -> dict[str, Any]:
    _ = start_line
    out = dict(metadata)
    heading_path = _markdown_heading_path_before(all_lines, end_line + 1)
    if heading_path:
        out["heading_path"] = heading_path
        out["section_title"] = heading_path[-1]
    return out


def _markdown_heading_path_before(lines: list[str], line_number: int) -> list[str]:
    path: list[str] = []
    for line in lines[: max(0, int(line_number) - 1)]:
        match = re.match(r"^\s*(#{1,6})\s+(.+?)\s*$", str(line or ""))
        if not match:
            continue
        level = len(match.group(1))
        title = re.sub(r"\s+#*$", "", match.group(2)).strip()
        if not title:
            continue
        path = path[: level - 1]
        path.append(title)
    return path


def _split_text(text: str, *, max_chars: int) -> Iterable[str]:
    normalized = re.sub(r"\n{3,}", "\n\n", str(text or "").strip())
    while len(normalized) > max_chars:
        split_at = normalized.rfind("\n", 0, max_chars)
        if split_at < 400:
            split_at = max_chars
        yield normalized[:split_at].strip()
        normalized = normalized[split_at:].strip()
    if normalized:
        yield normalized


def _merged_cell_values(sheet: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    ranges = getattr(getattr(sheet, "merged_cells", None), "ranges", []) or []
    for merged in ranges:
        top_value = sheet.cell(int(merged.min_row), int(merged.min_col)).value
        if top_value is None:
            continue
        for row_index in range(int(merged.min_row), int(merged.max_row) + 1):
            for col_index in range(int(merged.min_col), int(merged.max_col) + 1):
                values[(row_index, col_index)] = top_value
    return values


def _cell_value(sheet: Any, row_index: int, col_index: int, merged_values: dict[tuple[int, int], Any]) -> Any:
    value = sheet.cell(row_index, col_index).value
    if value is None:
        return merged_values.get((row_index, col_index))
    return value


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text[:_MAX_CELL_CHARS]
