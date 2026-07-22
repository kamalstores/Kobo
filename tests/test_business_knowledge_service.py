from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from opentulpa.business_knowledge.extraction import XLSX_MIME_TYPE, extract_source_sections
from opentulpa.business_knowledge.service import (
    BusinessKnowledgeService,
    OpenAICompatibleKnowledgeOracleClient,
)
from opentulpa.business_knowledge.table_normalizer import (
    select_table_evidence,
    table_evidence_selection_stats,
    table_evidence_to_toon,
    table_facts_from_sections,
)
from opentulpa.context.file_vault import FileVaultService


class _FakeOracle:
    def __init__(self, answer: str = "") -> None:
        self.answer_text = answer
        self.calls: list[dict[str, Any]] = []
        self.intent_calls: list[dict[str, Any]] = []

    def extract_intent(self, **kwargs: Any) -> dict[str, Any]:
        self.intent_calls.append(kwargs)
        query = str(kwargs.get("query", ""))
        if "what services" in query.casefold():
            return {
                "mode": "corpus_overview",
                "target_terms": [],
                "qualifier_terms": [],
                "ignore_terms": [],
            }
        if "ceramic" in query.casefold():
            return {
                "mode": "specific_fact",
                "target_terms": ["Ceramic wash"],
                "qualifier_terms": ["SUV", "Price"],
                "ignore_terms": [],
            }
        return {"mode": "specific_fact", "target_terms": [query], "qualifier_terms": [], "ignore_terms": []}

    def answer(self, **kwargs: Any) -> str:
        self.calls.append(kwargs)
        if self.answer_text:
            return self.answer_text
        source_pack = str(kwargs.get("source_pack", ""))
        query = str(kwargs.get("query", ""))
        if "reference" in query.casefold():
            return "Reference numbers are required for all appointments."
        if "unsupported" in query.casefold():
            return "NO_SOURCE"
        if "mode: overview" in source_pack and "Interior chemical cleaning" in source_pack:
            return "Available services include Ceramic wash and Interior chemical cleaning."
        if "Ceramic wash" in source_pack and "129" in source_pack:
            return "Ceramic wash for SUV costs 129. Source: prices.xlsx / Services row 2."
        return "NO_SOURCE"


def _vault(tmp_path: Path) -> FileVaultService:
    return FileVaultService(
        root_dir=tmp_path / "file_vault",
        db_path=tmp_path / "file_vault.db",
    )


def _knowledge(
    tmp_path: Path,
    vault: FileVaultService,
    *,
    oracle: _FakeOracle | None = None,
    max_source_pack_chars: int = 800_000,
) -> BusinessKnowledgeService:
    return BusinessKnowledgeService(
        root_dir=tmp_path / "knowledge",
        db_path=tmp_path / "knowledge.db",
        file_vault=vault,
        oracle_client=oracle or _FakeOracle(),
        max_source_pack_chars=max_source_pack_chars,
    )


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Services"
    ws.append(["Service", "Vehicle", "Price", "Notes"])
    ws.append(["Ceramic wash", "SUV", "129", "Exterior detail tier"])
    ws.append(["Interior chemical cleaning", "Sedan", "210", "Seats and carpets"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _dash_table_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tires"
    ws.append(["Size", "Sedan", "SUV", "Truck"])
    ws.append(["13R and smaller", "2500", "-----", "-----"])
    ws.append(["19R", "3000", "3500", "4000"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _duplicate_header_groups_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wash"
    ws.append(
        [
            "Service",
            "Class 1",
            "Class 1",
            "Class 2",
            "Class 2",
            "Class 3",
            "Class 3",
        ]
    )
    ws.append(["2-phase wash", "1000", "1000", "1200", "1200", "1400", "1400"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _bundle_and_direct_service_xlsx_bytes() -> bytes:
    wb = Workbook()
    packages = wb.active
    packages.title = "Packages"
    packages.append(["Package", "SUV"])
    packages.append(["Premium bundle: Basic wash; Wax; Interior clean", "199"])

    services = wb.create_sheet("Services")
    services.append(["Service", "SUV"])
    services.append(["Basic wash", "29"])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _sections_from_xlsx(raw_bytes: bytes, *, filename: str = "prices.xlsx"):
    sections, warnings, source_kind = extract_source_sections(
        record={
            "id": "test_file",
            "original_filename": filename,
            "mime_type": XLSX_MIME_TYPE,
        },
        raw_bytes=raw_bytes,
    )
    assert warnings == []
    assert source_kind == "structured_table"
    return sections


def test_business_knowledge_indexes_xlsx_and_queries_oracle(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="prices.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        caption=None,
        raw_bytes=_xlsx_bytes(),
    )
    oracle = _FakeOracle()
    knowledge = _knowledge(tmp_path, vault, oracle=oracle)

    indexed = knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_1",
        file_ids=[str(record["id"])],
    )

    assert indexed["sources"][0]["status"] == "indexed"
    assert indexed["sources"][0]["source_kind"] == "structured_table"
    assert indexed["index"]["engine"] == "knowledge_oracle"

    result = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_1",
        query="SUV ceramic wash price",
    )

    assert result.ok is True
    assert result.answer.answer_extract == "Ceramic wash for SUV costs 129. Source: prices.xlsx / Services row 2."
    assert oracle.intent_calls == [{"query": "SUV ceramic wash price"}]
    assert oracle.calls[0]["max_output_tokens"] == 1000
    assert "evidence_rows" in oracle.calls[0]["source_pack"]
    assert "Ceramic wash" in oracle.calls[0]["source_pack"]
    assert "129" in oracle.calls[0]["source_pack"]
    assert result.diagnostics["source_pack"]["mode"] == "specific_fact"
    assert result.diagnostics["source_pack"]["selection"]["row_count"] == 2
    assert result.diagnostics["source_pack"]["selection"]["candidate_count"] >= 1
    assert result.diagnostics["timing_ms"]["source_pack_total"] >= 0

    promoted = knowledge.promote_scope(
        customer_id="telegram_123",
        source_scope_type="workflow_setup",
        source_scope_id="iwsetup_1",
        target_scope_type="intake_workflow",
        target_scope_id="iwf_1",
    )
    assert promoted["source_count"] == 1
    assert promoted["section_count"] >= 1
    assert promoted["index"]["engine"] == "knowledge_oracle"


def test_table_normalizer_emits_header_bound_cell_facts_for_generic_xlsx() -> None:
    facts = table_facts_from_sections(_sections_from_xlsx(_xlsx_bytes()))

    price_fact = next(
        fact
        for fact in facts
        if fact.item == "Ceramic wash" and fact.header == "Price" and fact.value == "129"
    )
    vehicle_fact = next(
        fact
        for fact in facts
        if fact.item == "Ceramic wash" and fact.header == "Vehicle" and fact.value == "SUV"
    )

    assert price_fact.table == "Services"
    assert price_fact.column == "C"
    assert price_fact.value_kind == "number"
    assert vehicle_fact.column == "B"

    rows = select_table_evidence(
        facts,
        query="SUV ceramic wash price",
        target_terms=["Ceramic wash"],
        qualifier_terms=["SUV", "Price"],
    )
    toon = table_evidence_to_toon(
        rows,
        query="SUV ceramic wash price",
        target_terms=["Ceramic wash"],
        qualifier_terms=["SUV", "Price"],
    )

    assert rows[0].item == "Ceramic wash"
    assert "evidence_rows[rank,score,file,table,row,item,row_label,cells]:" in toon
    assert "Vehicle = SUV" in toon
    assert "Price = 129" in toon


def test_table_normalizer_keeps_fuzzy_typo_recall_with_bounded_candidates() -> None:
    facts = table_facts_from_sections(_sections_from_xlsx(_xlsx_bytes()))

    rows = select_table_evidence(
        facts,
        query="SUV ceramc wash price",
        target_terms=["ceramc wash"],
        qualifier_terms=["SUV", "Price"],
    )
    stats = table_evidence_selection_stats(
        facts,
        query="SUV ceramc wash price",
        target_terms=["ceramc wash"],
        qualifier_terms=["SUV", "Price"],
    )

    assert rows[0].item == "Ceramic wash"
    assert stats["candidate_count"] <= stats["candidate_limit"]
    assert stats["candidate_count"] >= 1


def test_table_normalizer_keeps_dash_only_cells_as_values_not_headers() -> None:
    facts = table_facts_from_sections(_sections_from_xlsx(_dash_table_xlsx_bytes()))

    row_facts = [fact for fact in facts if fact.item == "13R and smaller"]

    assert row_facts
    assert any(fact.header == "Sedan" and fact.value == "2500" for fact in row_facts)
    assert any(fact.header == "SUV" and fact.value == "-----" for fact in row_facts)
    assert any(fact.value_kind == "empty_marker" for fact in row_facts)

    rows = select_table_evidence(
        facts,
        query="13R and smaller sedan price",
        target_terms=["13R and smaller"],
        qualifier_terms=["Sedan"],
    )

    assert rows[0].item == "13R and smaller"


def test_table_normalizer_labels_repeated_header_groups_left_to_right() -> None:
    facts = table_facts_from_sections(_sections_from_xlsx(_duplicate_header_groups_xlsx_bytes()))

    rows = select_table_evidence(
        facts,
        query="2-phase wash class 3 price",
        target_terms=["2-phase wash"],
        qualifier_terms=["class 3"],
    )
    toon = table_evidence_to_toon(
        rows,
        query="2-phase wash class 3 price",
        target_terms=["2-phase wash"],
        qualifier_terms=["class 3"],
    )

    assert rows[0].item == "2-phase wash"
    assert "header_group 1 Class 1 [1] = 1000" in toon
    assert "header_group 3 Class 3 [1] = 1400" in toon


def test_table_normalizer_prefers_direct_service_rows_over_bundle_rows() -> None:
    facts = table_facts_from_sections(_sections_from_xlsx(_bundle_and_direct_service_xlsx_bytes()))

    rows = select_table_evidence(
        facts,
        query="SUV basic wash price",
        target_terms=["Basic wash"],
        qualifier_terms=["SUV"],
    )

    assert rows[0].table == "Services"
    assert rows[0].item == "Basic wash"
    assert any(row.table == "Packages" for row in rows[1:])


def test_business_knowledge_uses_overview_evidence_for_broad_table_questions(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="prices.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        caption=None,
        raw_bytes=_xlsx_bytes(),
    )
    oracle = _FakeOracle()
    knowledge = _knowledge(tmp_path, vault, oracle=oracle)
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_overview",
        file_ids=[str(record["id"])],
    )

    result = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_overview",
        query="what services do you offer?",
    )

    assert result.ok is True
    assert "Available services" in result.answer.answer_extract
    assert "mode: overview" in oracle.calls[0]["source_pack"]
    assert "tables[rank,table,row_count,sample_items]:" in oracle.calls[0]["source_pack"]
    assert "Interior chemical cleaning" in oracle.calls[0]["source_pack"]


def test_oracle_client_posts_default_model_with_openrouter_attribution(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    captured: dict[str, Any] = {}

    class _Response:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, Any]:
            return {"choices": [{"message": {"content": "grounded answer"}}]}

    def _post(url: str, **kwargs: Any) -> _Response:
        captured["url"] = url
        captured.update(kwargs)
        return _Response()

    monkeypatch.setattr("opentulpa.business_knowledge.oracle_client.httpx.post", _post)

    client = OpenAICompatibleKnowledgeOracleClient(
        api_key="test-key",
        base_url="https://openrouter.ai/api/v1",
        trace_path=tmp_path / "llm_call_traces.jsonl",
    )

    answer = client.answer(source_pack="SECRET RAW SOURCE CONTENT", query="price?")

    assert answer == "grounded answer"
    assert captured["url"] == "https://openrouter.ai/api/v1/chat/completions"
    assert captured["json"]["model"] == "google/gemini-3.1-flash-lite-preview"
    assert captured["json"]["max_tokens"] == 1000
    assert captured["json"]["reasoning"] == {"effort": "none", "exclude": True}
    assert captured["headers"]["Authorization"] == "Bearer test-key"
    assert captured["headers"]["HTTP-Referer"] == "https://github.com/kvyb/opentulpa"
    assert captured["headers"]["X-OpenRouter-Title"] == "OpenTulpa"
    trace_text = (tmp_path / "llm_call_traces.jsonl").read_text(encoding="utf-8")
    assert '"model_name": "google/gemini-3.1-flash-lite-preview"' in trace_text
    assert '"call_site": "knowledge_oracle"' in trace_text
    assert "SOURCE_PACK_SHA256" in trace_text
    assert "SECRET RAW SOURCE CONTENT" not in trace_text


def test_oracle_client_traces_intent_extraction(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    class _Response:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, Any]:
            return {
                "choices": [
                    {
                        "message": {
                            "content": (
                                '{"mode":"specific_fact",'
                                '"target_terms":["Ceramic wash"],'
                                '"qualifier_terms":["SUV"],'
                                '"ignore_terms":[]}'
                            )
                        }
                    }
                ]
            }

    def _post(url: str, **kwargs: Any) -> _Response:
        return _Response()

    monkeypatch.setattr("opentulpa.business_knowledge.oracle_client.httpx.post", _post)

    client = OpenAICompatibleKnowledgeOracleClient(
        api_key="test-key",
        base_url="https://openrouter.ai/api/v1",
        trace_path=tmp_path / "llm_call_traces.jsonl",
    )

    intent = client.extract_intent(query="SUV ceramic wash price")

    assert intent["target_terms"] == ["Ceramic wash"]
    assert intent["qualifier_terms"] == ["SUV"]
    trace_text = (tmp_path / "llm_call_traces.jsonl").read_text(encoding="utf-8")
    assert '"call_site": "knowledge_oracle_intent"' in trace_text
    assert "SOURCE_PACK_CHARS: 0" in trace_text


def test_business_knowledge_treats_oracle_no_source_as_empty_answer(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="policy.txt",
        mime_type="text/plain",
        caption=None,
        raw_bytes=b"Reference numbers are required for all appointments.",
    )
    knowledge = _knowledge(tmp_path, vault, oracle=_FakeOracle("NO_SOURCE"))
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_no_source",
        file_ids=[str(record["id"])],
    )

    result = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_no_source",
        query="unsupported question",
    )

    assert result.ok is False
    assert result.answer.answer_extract == ""


def test_business_knowledge_requeries_oracle_after_reindex(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="policy.txt",
        mime_type="text/plain",
        caption=None,
        raw_bytes=b"Reference numbers are required for all appointments.",
    )
    oracle = _FakeOracle()
    knowledge = _knowledge(tmp_path, vault, oracle=oracle)
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_cache",
        file_ids=[str(record["id"])],
    )

    first = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_cache",
        query="reference policy",
    )
    second = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_cache",
        query="reference policy",
    )

    assert first.cached is False
    assert second.cached is False
    assert "Reference numbers" in first.answer.answer_extract
    assert "Reference numbers" in second.answer.answer_extract
    assert len(oracle.calls) == 2

    vault.set_ai_summary("telegram_123", str(record["id"]), "updated")
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_cache",
        file_ids=[str(record["id"])],
    )
    third = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_cache",
        query="reference policy",
    )

    assert third.cached is False
    assert "Reference numbers" in third.answer.answer_extract
    assert len(oracle.calls) == 3


def test_business_knowledge_preflight_reuses_durable_cache_for_same_source_goal(
    tmp_path: Path,
) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="policy.txt",
        mime_type="text/plain",
        caption=None,
        raw_bytes=b"Reference numbers are required for all appointments.",
    )
    oracle = _FakeOracle("The source supports appointment intake.")
    knowledge = _knowledge(tmp_path, vault, oracle=oracle)
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache",
        file_ids=[str(record["id"])],
    )

    first = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache",
        workflow_goal="appointment intake",
    )
    second = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache",
        workflow_goal="appointment intake",
    )

    assert first["ok"] is True
    assert first["cache_hit"] is False
    assert second["ok"] is True
    assert second["cache_hit"] is True
    assert second["diagnostics"]["cache"]["hit"] is True
    assert len(oracle.calls) == 1

    resumed_oracle = _FakeOracle("should not be called")
    resumed = BusinessKnowledgeService(
        root_dir=tmp_path / "knowledge",
        db_path=tmp_path / "knowledge.db",
        file_vault=vault,
        oracle_client=resumed_oracle,
    )
    cached_after_restart = resumed.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache",
        workflow_goal="appointment intake",
    )

    assert cached_after_restart["cache_hit"] is True
    assert resumed_oracle.calls == []


def test_business_knowledge_preflight_cache_invalidates_on_goal_or_source_change(
    tmp_path: Path,
) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="policy.txt",
        mime_type="text/plain",
        caption=None,
        raw_bytes=b"Reference numbers are required for all appointments.",
    )
    oracle = _FakeOracle("The source supports appointment intake.")
    knowledge = _knowledge(tmp_path, vault, oracle=oracle)
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache_invalidate",
        file_ids=[str(record["id"])],
    )

    first = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache_invalidate",
        workflow_goal="appointment intake",
    )
    changed_goal = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache_invalidate",
        workflow_goal="appointment intake with vehicle class",
    )
    vault.set_ai_summary("telegram_123", str(record["id"]), "summary changed")
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache_invalidate",
        file_ids=[str(record["id"])],
    )
    changed_source = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_preflight_cache_invalidate",
        workflow_goal="appointment intake",
    )

    assert first["cache_hit"] is False
    assert changed_goal["cache_hit"] is False
    assert changed_source["cache_hit"] is False
    assert len(oracle.calls) == 3


def test_business_knowledge_fails_when_source_pack_is_too_large(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="policy.txt",
        mime_type="text/plain",
        caption=None,
        raw_bytes=b"Reference numbers are required for all appointments.",
    )
    oracle = _FakeOracle()
    knowledge = _knowledge(tmp_path, vault, oracle=oracle, max_source_pack_chars=20)
    knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_limit",
        file_ids=[str(record["id"])],
    )

    result = knowledge.query(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_limit",
        query="reference policy",
    )

    assert result.ok is False
    assert result.answer.answer_extract == ""
    assert "exceeds" in " ".join(result.warnings)
    assert oracle.calls == []


def test_business_knowledge_configures_sqlite_for_concurrent_server_use(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    knowledge = _knowledge(tmp_path, vault)

    with knowledge.repository.conn() as conn:
        journal_mode = str(conn.execute("PRAGMA journal_mode").fetchone()[0]).lower()
        busy_timeout = int(conn.execute("PRAGMA busy_timeout").fetchone()[0])
        synchronous = int(conn.execute("PRAGMA synchronous").fetchone()[0])

    assert journal_mode == "wal"
    assert busy_timeout >= 10_000
    assert synchronous == 1


def test_business_knowledge_flags_unsupported_binary_as_not_grounded(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="menu.bin",
        mime_type="application/octet-stream",
        caption=None,
        raw_bytes=b"\x00\x01\x02",
    )
    knowledge = _knowledge(tmp_path, vault)

    indexed = knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_unsupported",
        file_ids=[str(record["id"])],
    )
    preflight = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_unsupported",
        workflow_goal="service prices",
    )

    assert indexed["sources"][0]["status"] == "unsupported"
    assert preflight["ok"] is False
    assert preflight["status"] == "needs_better_source"
    assert preflight["diagnostics"]["timing_ms"]["preflight_total"] >= 0
    assert "unsupported" in " ".join(preflight["warnings"]).lower()


def test_business_knowledge_indexes_media_summary_as_derived_not_authoritative(tmp_path: Path) -> None:
    vault = _vault(tmp_path)
    record = vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="photo",
        telegram_file_id=None,
        original_filename="menu.png",
        mime_type="image/png",
        caption=None,
        raw_bytes=b"not really a png",
    )
    vault.set_ai_summary(
        "telegram_123",
        str(record["id"]),
        "The image appears to show a basic wash price of 10.",
    )
    knowledge = _knowledge(tmp_path, vault, oracle=_FakeOracle("The image summary mentions basic wash 10."))

    indexed = knowledge.index_sources(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_media",
        file_ids=[str(record["id"])],
    )
    preflight = knowledge.preflight_scope(
        customer_id="telegram_123",
        scope_type="workflow_setup",
        scope_id="iwsetup_media",
        workflow_goal="basic wash price",
    )

    assert indexed["sources"][0]["status"] == "indexed"
    assert indexed["sources"][0]["source_kind"] == "derived_from_media"
    assert preflight["ok"] is False
    assert preflight["status"] == "needs_better_source"
