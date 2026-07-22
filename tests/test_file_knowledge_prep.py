from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from opentulpa.agent.file_analysis import extract_uploaded_text
from opentulpa.agent.knowledge_prep import inspect_uploaded_file_structure
from opentulpa.api.app import create_app
from opentulpa.context.file_vault import FileVaultService


class _DisabledComposio:
    enabled = False

    def status(self) -> dict[str, object]:
        return {"ok": True, "enabled": False}


def _autospa_workbook_bytes() -> bytes:
    workbook = Workbook()
    wash = workbook.active
    wash.title = "Мойка"
    wash.append(["Услуга", "C-Class", "SUV"])
    wash.append(["2х-фазная мойка кузова", 1000, 1200])
    tire = workbook.create_sheet("Шиномонтаж")
    tire.append(
        [
            "Размерность дисков",
            "Седан",
            "Внедорожник / кросовер + низкий профиль",
        ]
    )
    tire.append(["Комплект 19R", 3000, 4000])
    ppf = workbook.create_sheet("PPF")
    ppf.append(["Пакет", "Цена"])
    ppf.append(["Передняя часть", 50000])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_xlsx_inspection_returns_structure_and_search_matches() -> None:
    inspected = inspect_uploaded_file_structure(
        raw_bytes=_autospa_workbook_bytes(),
        filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        search_terms=["Комплект", "Шиномонтаж"],
    )

    sheets = inspected["structure"]["sheets"]
    assert [sheet["name"] for sheet in sheets] == ["Мойка", "Шиномонтаж", "PPF"]
    tire_sheet = sheets[1]
    assert tire_sheet["matched_terms"] == ["Шиномонтаж"]
    assert tire_sheet["sample_rows"][0]["source_ref"] == "Шиномонтаж!1"
    assert tire_sheet["table_candidates"][0]["row_start"] == 1
    assert tire_sheet["matches"][0]["source_ref"] == "Шиномонтаж!2"


def test_xlsx_upload_text_extraction_returns_workbook_preview() -> None:
    extracted = extract_uploaded_text(
        raw_bytes=_autospa_workbook_bytes(),
        filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    assert "# Sheet 1: Мойка" in extracted
    assert "2х-фазная мойка кузова" in extracted
    assert "# Sheet 2: Шиномонтаж" in extracted
    assert "Комплект 19R" in extracted


def test_inspect_structure_route_returns_workbook_inventory(tmp_path: Path) -> None:
    file_vault = FileVaultService(
        root_dir=tmp_path / "file_vault",
        db_path=tmp_path / "file_vault.db",
    )
    source = file_vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        caption=None,
        raw_bytes=_autospa_workbook_bytes(),
    )
    app = create_app(
        file_vault_service=file_vault,
        composio_service=_DisabledComposio(),
    )

    with TestClient(app) as client:
        response = client.post(
            "/internal/files/inspect_structure",
            json={
                "customer_id": "telegram_123",
                "file_id": source["id"],
                "search_terms": ["низкий профиль", "Шиномонтаж"],
            },
        )

    assert response.status_code == 200
    inspection = response.json()["inspection"]
    assert inspection["format"] == "xlsx"
    tire_sheet = inspection["structure"]["sheets"][1]
    assert tire_sheet["name"] == "Шиномонтаж"
    assert tire_sheet["matched_terms"] == ["Шиномонтаж"]
    assert tire_sheet["matches"][0]["source_ref"] == "Шиномонтаж!1"
